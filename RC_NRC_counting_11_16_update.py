# Goal: 輸出工時數值     2024/11/16 update
# Input: MBOM-後半部除工時外的其他欄位 
# 完工者: 逢大專題生(2024畢)

import pandas as pd
from openpyxl import load_workbook
import numpy as np
import re

no = []
widthin = []
lengthin = []
material = []
spec = []
outsourcing = []
process_code = []
thickin = []

VA = [None]*8000
OSM = [None]*8000
NCTP_S = [None]*8000
NCTP_W = [None]*8000
NCTP_D = [None]*8000
NCMF = [None]*8000
NCTP_E = [None]*8000
PST = [None]*8000
HFB = [None]*8000
TRH_T = [None]*8000
SFS = [None]*8000
NCTP_R = [None]*8000
HOFX = [None]*8000
TCT = [None]*8000
MST = [None]*8000
SFB = [None]*8000
CMT = [None]*8000
RC = [None]*8000

def read_file(): # 讀取「最終報價工時BOM.xlsx」並將各欄位每一行的資料儲存，以方便後續計算
    df = pd.read_excel('最終報價工時BOM.xlsx', sheet_name='REF MAKE', na_values=['nan'])

    global no
    no = df['NO'][0:].values

    global widthin
    widthin = df['寬度\nWidth\nin'][0:].values

    global lengthin
    lengthin = df['長度\nLength\nin'][0:].values

    global material
    material = df['材質\nMaterial'][0:].values

    global spec
    # spec = df['擠製件/蜂巢件規範Spec'][0:].values
    spec = df['擠製件/蜂巢件規範Spec'].values.astype(str)

    global outsourcing
    outsourcing = df['外包件\n(Outsourcing)'][0:].values
    for i in range(len(no)):
        string = str(outsourcing[i])
        if string.startswith('OPP'):
            outsourcing[i] = re.sub(r'\d+', '', string)

    global process_code
    process_code = df['製程分類'][0:].values

    global thickin
    thickin = df['厚度\nThickness\nin'][0:].values

    # 至少有 20 筆資料才算載入成功
    count = 0
    for i in range(len(no)):
        if not np.isnan(lengthin[i]):
            count += 1
        if count > 20:
            break
    if count > 20:
        print('載入成功\n執行中...')
    else:
        print('載入失敗，請先手動儲存Excel後再執行程式')
        exit()
            
def function(): # 將計算 NRC 和 RC 的各函式進行整合
    for i in range(len(no)):
        polished(i)
        SB(i)
        SB_handwork(i)
        SH(i)
        SC(i)
        SS(i)
        SS_handwork(i)
        SSC(i)
        SR(i)
        SSCM(i)
        SM(i)
        SSEM(i)
        SSM(i)
        SB_RC(i)
        SR_RC(i)
        SS_RC(i)
        SH_RC(i)
        SM_RC(i)
        SME_RC(i)
        SC_RC(i)
        SSC_RC(i)
        SSCM_RC(i)
        SSM_RC(i)
        SSEM_RC(i)        

def save_file(): # 將計算完的工時資料，寫入原 excel 的對應的欄位
    global NCTP_D
    global no
    global NCMF
    # 讀取 Excel 檔案
    wb = load_workbook('最終報價工時BOM.xlsx')

    # 選擇特定的資料表
    ws = wb['REF MAKE']

    # 將陣列的資料填入特定的列
    for i in range(0, len(no)):
        # 將拋光存入 excel 中的外包件
        ws.cell(row = i+2, column = 27).value = outsourcing[i]
        
        # NRC、RC 工時填回 excel 中
        if OSM[i] != None:
            ws.cell(row = i+3, column = 47).value = OSM[i]
        else:
            ws.cell(row = i+3, column = 47).value = ''
        if NCTP_S[i] != None:
            ws.cell(row = i+3, column = 48).value = NCTP_S[i]
        else:
            ws.cell(row = i+3, column = 48).value = ''
        if NCTP_W[i] != None:
            ws.cell(row = i+3, column = 49).value = NCTP_W[i]
        else:
            ws.cell(row = i+3, column = 49).value = ''
        if NCTP_E[i] != None:
            ws.cell(row = i+3, column = 50).value = NCTP_E[i]
        else:
            ws.cell(row = i+3, column = 50).value = ''
        if VA[i] != None:
            ws.cell(row = i+3, column = 51).value = VA[i]
        else:
            ws.cell(row = i+3, column = 51).value = ''
        if CMT[i] != None:
            ws.cell(row = i+3, column = 52).value = CMT[i]
        else:
            ws.cell(row = i+3, column = 52).value = CMT[i]
        if SFB[i] != None:
            ws.cell(row = i+3, column = 53).value = SFB[i]
        else:
            ws.cell(row = i+3, column = 53).value = ''
        if HFB[i] != None:
            ws.cell(row = i+3, column = 54).value = HFB[i]
        else:
            ws.cell(row = i+3, column = 54).value = ''
        if TCT[i] != None:
            ws.cell(row = i+3, column = 55).value = TCT[i]
        else:
            ws.cell(row = i+3, column = 55).value = ''
        if TRH_T[i] != None:
            ws.cell(row = i+3, column = 56).value = TRH_T[i]
        else:
            ws.cell(row = i+3, column = 56).value = ''
        if SFS[i] != None:
            ws.cell(row = i+3, column = 57).value = SFS[i]
        else:
             ws.cell(row = i+3, column = 57).value = ''
        if NCTP_D[i] != None:
            ws.cell(row = i+3, column = 58).value = NCTP_D[i]
        else:
            ws.cell(row = i+3, column = 58).value = ''
        if NCMF[i] != None:
            ws.cell(row = i+3, column = 59).value = NCMF[i]
        else:
            ws.cell(row = i+3, column = 59).value = ''
        if NCTP_R[i] != None:
            ws.cell(row = i+3, column = 60).value = NCTP_R[i]
        else:
            ws.cell(row = i+3, column = 60).value = ''
        if HOFX[i] != None:
            ws.cell(row = i+3, column = 61).value = HOFX[i]
        else:
            ws.cell(row = i+3, column = 61).value = ''
        if MST[i] != None:
            ws.cell(row = i+3, column = 62).value = MST[i]
        else:
            ws.cell(row = i+3, column = 62).value = ''
        if PST[i] != None:
            ws.cell(row = i+3, column = 63).value = PST[i]
        else:
            ws.cell(row = i+3, column = 63).value = ''
        if RC[i] != None:
            ws.cell(row = i+3, column = 64).value = RC[i]
        else:
            ws.cell(row = i+3, column = 64).value = ''

    # 儲存 Excel 檔案
    wb.save('最終報價工時BOM.xlsx')
    print('執行完畢')

def polished(i): # 判斷是否須加上「拋光」
        string = str(material[i])
        if 'POLISHED' in string:
            # 有則在外包件加上「拋光」
            outsourcing[i] = str(outsourcing[i]) + '/拋光'

def SB(i):  # SB的NRC計算
    if process_code[i] == 'SB':
        length = 0
        width = 0
        string = str(outsourcing[i])
        if '五面' in string:
            length = 0
            width = 0
            if not np.isnan(lengthin[i]):
                length = lengthin[i]
            if not np.isnan(widthin[i]):
                width = widthin[i]
            a = max(width, length)  # 取最大值
            if a < 16:
                NCTP_D[i-1] = 8  # 直接對列表元素賦值，不使用insert
                NCMF[i-1] = 72
            else:
                NCTP_D[i-1] = 10
                NCMF[i-1] = 96
        if '摺角度' in string:
            length = 0
            width = 0
            if not np.isnan(lengthin[i]):
                length = lengthin[i]
            if not np.isnan(widthin[i]):
                width = widthin[i]
            a = max(width, length)
            if a < 21:
                NCTP_E[i-1] = 6
            elif 21 <= a < 31:
                NCTP_E[i-1] = 8
            else:
                NCTP_E[i-1] = 12
        if 'EB' in string:
            length = 0
            width = 0
            if not np.isnan(lengthin[i]):
                length = lengthin[i]
            if not np.isnan(widthin[i]):
                width = widthin[i]
            a = max(width, length)
            if a < 16:
                PST[i-1] = 20
            elif 16 <= a < 31:
                PST[i-1] = 32
            else:
                PST[i-1] = 40
        if 'OPP' in string:
            OSM[i-1] = 0
            NCTP_S[i-1] = 0
            VA[i-1] = 0
        else:
            OSM[i-1] = 6
            VA[i-1] = 2
            if not '摺角度' in string:
                NCTP_S[i-1] = 6
            else:
                NCTP_S[i-1] = None

def SB_handwork(i): # SB手工的NRC計算
    if process_code[i] == 'SB+手工':
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length)
        if a < 21:
            NCTP_S[i-1] = 6
            OSM[i-1] = 6
            VA[i-1] = 2
        elif 21 <= a < 31:
            NCTP_S[i-1] = 6
            OSM[i-1] = 6
            VA[i-1] = 2
        else:
            NCTP_S[i-1] = 6
            OSM[i-1] = 6
            VA[i-1] = 2
        if a < 16:
            HFB[i-1] = 32
        elif 16 <= a < 31:
            HFB[i-1] = 48
        else:
            HFB[i-1] = 70
                
def SH(i):  # SH的NRC計算
    if process_code[i] == 'SH':
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]

        a = max(length, width)
        if a >= 16: #外放扣4
            a -= 4

        string = str(outsourcing[i])

        if '五面' in string:
            if a < 16:
                # NCTP_S
                NCTP_D[i-1] = 8  # 直接對列表元素賦值，不使用insert
                NCMF[i-1] = 72
            else :
                NCTP_D[i-1] = 10
                NCMF[i-1] = 96

        if '手臂' in string:
            if a < 21:
                NCTP_R[i - 1] = 12
                HOFX[i - 1] = 150
            elif 21 <= a < 51:
                NCTP_R[i - 1] = 16
                HOFX[i - 1] = 300
            elif 51 <= a < 70 :
                NCTP_R[i - 1] = 20
                HOFX[i - 1] = 520
            else :
                NCTP_R[i - 1] = 20
                HOFX[i - 1] = 700
        if '2道模' in string:
            if a < 16:
                SFB[i - 1] = 52 #20*2 + 12 加工工時2倍
            elif 16 <= a < 31:
                SFB[i - 1] = 80
            elif 31 <= a < 51:
                SFB[i - 1] = 120
            else :
                SFB[i - 1] = 190
        else :
            if a < 16:
                SFB[i - 1] = 32
            elif 16 <= a < 31:
                SFB[i - 1] = 48
            elif 31 <= a < 51:
                SFB[i - 1] = 70
            else :
                SFB[i - 1] = 110

        b = a % 10
        a = a / 10
        c = int(a)
        if b == 0:
            NCTP_S[i-1] = 8+(c-1)*2
        else:
            NCTP_S[i-1] = 8+c*2

        OSM[i-1] = 6
        
def SC(i):  # SC的NRC計算
    if process_code[i] == 'SC':
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(length, width) #外放
        string = str(outsourcing[i])
        
        if a < 21:
            CMT[i-1] = 52
        elif 21 <= a < 41:
            CMT[i-1] = 72
        else:
            CMT[i-1] = 96

        if a < 21:
            NCTP_S[i-1] = 4 * 2
        elif 21 <= a < 31:
            NCTP_S[i-1] = 6 * 2
        else:
            NCTP_S[i-1] = 8 * 2

        OSM[i-1] = 6

def SS(i):  # SS的NRC計算
    if process_code[i] == 'SS':
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(length, width) #外放
        if a > 100:
            a -= 35
        string = str(outsourcing[i])

        if '五面' in string:
            if a < 16:
                NCTP_D[i-1] = 8  #
                NCMF[i-1] = 72
            else :
                NCTP_D[i-1] = 10
                NCMF[i-1] = 96
        if '手臂' in string:
            if a < 21:
                NCTP_R[i - 1] = 12
                HOFX[i - 1] = 150
            elif 21 <= a < 51:
                NCTP_R[i - 1] = 16
                HOFX[i - 1] = 300
            elif 51 <= a < 70 :
                NCTP_R[i - 1] = 20
                HOFX[i - 1] = 520
            else :
                NCTP_R[i - 1] = 20
                HOFX[i - 1] = 700
        
        if a < 41:
            TRH_T[i-1] = 46
        else:
            TRH_T[i-1] = 64

        if a < 31:
            SFS[i-1] = 170
        elif 31 <= a < 51:
            SFS[i-1] = 235
        elif 51 <= a < 101:
            SFS[i-1] = 410
        elif 101 <= a < 151:
            SFS[i-1] = 545
        elif 151 <= a < 201:
            SFS[i-1] = 710
        else:
            SFS[i-1] = 800
        
        OSM[i-1] = 6
            
def SSC(i): # SSC的NRC計算
    if process_code[i] == 'SSC':
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = CMTN = TRHN = SFSN = NCTPRN = HOFXN = MSTN = max(length, width)

        string = str(outsourcing[i])

        if '五面' in string:
            if a < 16:
                NCTP_D[i-1] = 8  #
                NCMF[i-1] = 72
            else :
                NCTP_D[i-1] = 10
                NCMF[i-1] = 96
        if 125 > NCTPRN >= 110:
            NCTPRN -= 75
        elif 110 >= NCTPRN > 70:
            NCTPRN -= 55
        elif 70 >= NCTPRN > 50:   
            NCTPRN -= 42

        if HOFXN > 110:
            HOFXN -= 55
        elif 110 >= HOFXN > 70:
            HOFXN -= 35
        elif 70 >= HOFXN > 50:
            HOFXN -= 12
        if '手臂' in string:
            if NCTPRN < 21:
                NCTP_R[i - 1] = 12
            elif 21 <= NCTPRN < 51:
                NCTP_R[i - 1] = 16
            elif 51 <= NCTPRN < 70 :
                NCTP_R[i - 1] = 20
            else :
                NCTP_R[i - 1] = 20
            
            if HOFXN < 21:
                HOFX[i - 1] = 300 #??? 150 ???
            elif 21 <= HOFXN < 51:
                HOFX[i - 1] = 300
            elif 51 <= HOFXN < 70 :
                HOFX[i - 1] = 520
            else :
                HOFX[i - 1] = 700

        MSTN -= 6
        if '拋光' in string:
            if MSTN < 21:
                MST[i-1] = 84
            elif 21 <= MSTN < 41:
                MST[i-1] = 120
            else :
                MST[i-1] = 156
        else:
            MST[i-1] = 0
        
        if CMTN > 50:
            CMTN -= 22
        else:
            CMTN -= 6

        if CMTN < 21:
            CMT[i-1] = 52
        elif 21 <= CMTN < 41:
            CMT[i-1] = 72
        else:
            CMT[i-1] = 96

        TRHN -= 6
        if TRHN < 41:
            TRH_T[i-1] = 46
        else:
            TRH_T[i-1] = 64

        SFSN -= 26
        if SFSN < 31:
            SFS[i-1] = 170
        elif 31 <= SFSN < 51:
            SFS[i-1] = 235
        elif 51 <= SFSN < 101:
            SFS[i-1] = 410
        elif 101 <= SFSN < 151:
            SFS[i-1] = 545
        elif 151 <= SFSN < 201:
            SFS[i-1] = 710
        else:
            SFS[i-1] = 800
        
        OSM[i-1] = 6

def SR(i):  # SR的NRC計算
    if process_code[i] == 'SR':
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(length, width) #外放-3~4
        string = str(outsourcing[i])

        if '五面' in string:
            if a < 16:
                NCTP_D[i-1] = 8  #
                NCMF[i-1] = 72
            else :
                NCTP_D[i-1] = 10
                NCMF[i-1] = 96

        if a < 19: #外放3~4 原16
            TCT[i-1] = 32
        elif 19 <= a < 34:
            TCT[i-1] = 40
        else:
            TCT[i-1] = 48
        
        OSM[i-1] = 6
        NCTP_S[i-1] = 6

def SS_handwork(i): # SS手工的NRC計算
    if process_code[i] == 'SS+手工':
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length)
        
        if a < 16:
            HFB[i-1] = 32
            OSM[i-1] = 6
            NCTP_S[i-1] = 6
        elif 16 <= a < 31:
            HFB[i-1] = 48
            OSM[i-1] = 6
            NCTP_S[i-1] = 6
        else:
            HFB[i-1] = 70
            OSM[i-1] = 6
            NCTP_S[i-1] = 6
            
        if a < 21:
            TRH_T[i-1] = 36
        elif 21 <= a < 41:
            TRH_T[i-1] = 46
        else:
            TRH_T[i-1] = 64
            
        if a < 21:
            NCTP_R[i-1] = 12
            HOFX[i-1] = 150
        elif 21 <= a < 51:
            NCTP_R[i-1] = 16
            HOFX[i-1] = 300
        else:
            NCTP_R[i-1] = 20
            HOFX[i-1] = 520
            
        if a < 31:
            SFS[i-1] = 170
        elif 31 <= a < 51:
            SFS[i-1] = 235
        elif 51 <= a < 101:
            SFS[i-1] = 410
        elif 101 <= a < 151:
            SFS[i-1] = 545
        elif 151 <= a < 201:
            SFS[i-1] = 710
        else:
            SFS[i-1] = 800
         
def SSCM(i):    # SSCM的NRC計算
    if process_code[i] == 'SSCM':
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length) - 5 #外放
        OSM[i-1] = 6

        if a < 21:
            CMT[i-1] = 52
        elif 21 <= a < 41:
            CMT[i-1] = 72
        else:
            CMT[i-1] = 96

        if a < 21:
            TRH_T[i-1] = 36
        elif 21 <= a < 41:
            TRH_T[i-1] = 46
        else:
            TRH_T[i-1] = 64

        if a < 31:
            SFS[i-1] = 170
        elif 31 <= a < 51:
            SFS[i-1] = 235
        elif 51 <= a < 101:
            SFS[i-1] = 410
        elif 101 <= a < 151:
            SFS[i-1] = 545
        elif 151 <= a < 201:
            SFS[i-1] = 710
        else:
            SFS[i-1] = 800
            
        MST[i-1] = 120 #目前設定固定120
         
def SM(i):  # SM的NRC計算
    if process_code[i] == 'SM':
        length = 0
        width = 0
        string = str(outsourcing[i])
        if 'OPP' in string:
            OSM[i-1] = 0
            NCTP_E[i-1] = 0
        else:
            OSM[i-1] = 6
        if spec[i] != 'nan' and 'OPP' not in string:
            NCTP_E[i-1] = 6
        else:
            length = 0
            width = 0
            if not np.isnan(lengthin[i]):
                length = lengthin[i]
            if not np.isnan(widthin[i]):
                width = widthin[i]
            a = max(width, length)
            if 'CRES' in material[i]:
                if a < 21:
                    NCTP_W[i-1] = 4
                elif 21 <= a < 31:
                    NCTP_W[i-1] = 6
                else:
                    NCTP_W[i-1] = 8
            else:
                if not 'OPP' in string:
                    if a < 21:
                        NCTP_S[i-1] = 4
                    elif 21 <= a < 31:
                        NCTP_S[i-1] = 6
                    else:
                        NCTP_S[i-1] = 8
        if '倒角' in string:
            VA[i-1] = 2
        if '五面' in string:
            length = 0
            width = 0
            if not np.isnan(lengthin[i]):
                length = lengthin[i]
            if not np.isnan(widthin[i]):
                width = widthin[i]
            a = max(width, length)
            if a < 16:
                NCTP_D[i-1] = 8
                NCMF[i-1] = 72
            else:
                NCTP_D[i-1] = 10
                NCMF[i-1] = 96
                    
def SSEM(i):    # SSEM的NRC計算
    if process_code[i] == 'SSEM':
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length)
        if a > 100:
            a -=25
        if a < 31:
            SFS[i-1] = 170
            OSM[i-1] = 6
        elif 31 <= a < 51:
            OSM[i-1] = 6
            SFS[i-1] = 235
        elif 51 <= a < 101:
            OSM[i-1] = 6
            SFS[i-1] = 410
        elif 101 <= a < 151:
            OSM[i-1] = 6
            SFS[i-1] = 545
        elif 151 <= a < 201:
            OSM[i-1] = 6
            SFS[i-1] = 710
        else:
            OSM[i-1] = 6
            SFS[i-1] = 800
                
def SSM(i): # SSM的NRC計算
    if process_code[i] == 'SSM':
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length)
        if a > 100:
            a -= 25
        if a < 31:
            SFS[i-1] = 170
            OSM[i-1] = 6
        elif 31 <= a < 51:
            OSM[i-1] = 6
            SFS[i-1] = 235
        elif 51 <= a < 101:
            OSM[i-1] = 6
            SFS[i-1] = 410
        elif 101 <= a < 151:
            OSM[i-1] = 6
            SFS[i-1] = 545
        elif 151 <= a < 201:
            OSM[i-1] = 6
            SFS[i-1] = 710
        else:
            OSM[i-1] = 6
            SFS[i-1] = 800
        
        if a < 21:
            TRH_T[i-1] = 36
        elif 21 <= a < 41:
            TRH_T[i-1] = 46
        else:
            TRH_T[i-1] = 64
    
def SB_RC(i):   # SB的RC計算
    if process_code[i] == 'SB' or process_code[i] == 'SB+手工' :
        time = 0
        length = 0
        width = 0
        thick = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        if not np.isnan(thickin[i]):
            thick = thickin[i]
        a = max(width, length)

        if a > 10: #外放
            a -= 2

        if a < 11:
            time += 0.25
        elif 11 <= a < 21:
            time += 0.27
        elif 21 <= a < 31:
            time += 0.3
        elif 31 <= a < 41:
            time += 0.4
        elif 41 <= a < 51:
            time += 0.5
        elif 51 <= a < 61:
            time += 0.6
        elif 61 <= a < 71:
            time += 0.7
        elif 71 <= a < 81:
            time += 0.8
        elif 81 <= a < 91:
            time += 0.9
        else:
            time += 1
        string = str(outsourcing[i])
        if '道' in string:
            n = r'(\d+)道'
            match = re.match(n, string)
            if match:
                number = int(match.group(1))
                number -= 1
            time += number*0.08
        if '五面' in string:
            time += 0.3
        if '劃切' in string:
            time += 0.2
        if 'EB' in string:
            time += 0.3
        
        string2 = str(process_code[i]) 
        if'手工' in string2 :
            if a < 11:
                time += 0.5
            elif 11 <= a < 21:
                time += 0.7
            elif 21 <= a < 31:
                time += 0.9
            elif 31 <= a < 41:
                time += 1.2
            elif 41 <= a < 51:
                time += 2
            elif 51 <= a < 61:
                time += 2.5
            elif 61 <= a < 71:
                time += 3
            elif 71 <= a < 81:
                time += 3.5
            elif 81 <= a < 91:
                time += 4
            else :
                time += 5

        if '摺角度' in string:
            if a < 11:
                time += 0.25
            elif 11 <= a < 21:
                time += 0.27
            elif 21 <= a < 31:
                time += 0.3
            elif a <= 31 < 41:
                time += 0.35
            elif 41 <= a < 51:
                time += 0.4
            elif 51 <= a < 61:
                time += 0.45
            elif 61 <= a < 71:
                time += 0.5
            elif 71 <= a < 81:
                time += 0.55
            elif 81 <= a < 91:
                time += 0.6
            else:
                time += 0.7
        RC[i-1] = time
        if 'OPP' in string:
            for j in range(len(no)):
                if  process_code[j] == 'SB' or process_code[j] == 'SB+手工':
                    length2 = 0
                    width2 = 0
                    thick2 = 0
                    if not np.isnan(lengthin[j]):
                        length2 = lengthin[j]
                    if not np.isnan(widthin[j]):
                        width2 = widthin[j]
                    if not np.isnan(thickin[j]):
                        thick2 = thickin[j]
                    if thick2 == thick and width2 == width and length2 == length:
                        RC[i-1] = RC[j-1]
                        break

def SR_RC(i):   # SR的RC計算
    if process_code[i] == 'SR':
        time = 0
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length)   

        if a < 11:
            time += 0.5
        elif 11 <= a < 21:
            time += 0.7
        elif 21 <= a < 31:
            time += 0.9
        elif 31 <= a < 51:
            time += 1
        elif 51 <= a < 71:
            time += 1.2
        elif 71 <= a < 91:
            time += 1.5
        elif 91 <= a < 111:
            time += 2
        elif 111 <= a < 131:
            time += 3
        elif 131 <= a < 151:
            time += 4
        else:
            time += 5

        string = str(outsourcing[i])
        if '五面' in string:
            time += 0.3
        if '手臂' in string:
            time += 0.5 

        RC[i-1] = time

def SS_RC(i):   # SS的RC計算
    if process_code[i] == 'SS' or process_code[i] == 'SS+手工':
        time = 0
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length)
        if a > 50:
            a -= 30

        if a < 31:
            time += 2
        elif 31 <= a < 51:
            time += 2.5
        elif 51 <= a < 71:
            time += 3
        elif 71 <= a < 91:
            time += 4
        else :
            time += 5

        string = str(outsourcing[i])
        if '五面' in string:
            time += 0.3
        if '手臂' in string:
            time += 0.5

        string2 = str(process_code[i]) 
        if '手工' in string2 :
            time += 0.5   

        RC[i-1] = time

def SH_RC(i):   # SH的RC計算
    if process_code[i] == 'SH':
        time = 0
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length)
        if a > 50:  #外放值大約為5,1230不對
            a -= 5

        if a < 11:
            time += 0.5
        elif 11 <= a < 21:
            time += 0.7
        elif 21 <= a < 31:
            time += 0.9
        elif 31 <= a < 41:
            time += 1.2
        elif 41 <= a < 51:
            time += 2
        elif 51 <= a < 61:
            time += 2.5
        elif 61 <= a < 71:
            time += 3
        elif 71 <= a < 81:
            time += 3.5
        elif 81 <= a < 91:
            time += 4
        else :
            time += 5

        string = str(outsourcing[i])
        if '2道' in string:
            time += 0.5
        if '五面' in string:
            time += 0.3
        if '手臂' in string:
            time += 0.5

        RC[i-1] = time

def SM_RC(i):   # SM的RC計算
    if process_code[i] == 'SM' and spec[i] == 'nan':
        time = 0
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length) - 2 #外放
        if a < 11:
            time += 0.15
        elif 11 <= a < 21:
            time += 0.2
        elif 21 <= a < 31:
            time += 0.22
        elif 31 <= a < 41:
            time += 0.25
        elif 41 <= a < 51:
            time += 0.3
        elif 51 <= a < 61:
            time += 0.35
        elif 61 <= a < 71:
            time += 0.4
        elif 71 <= a < 81:
            time += 0.5
        elif 81 <= a < 91:
            time += 0.6
        else :
            time += 0.7

        string = str(outsourcing[i])
        if '五面' in string:
            time += 0.3
        if '手臂' in string:
            time += 0.5
        if '倒角' in string:
            time += 0.2
        if '劃切' in string:
            time += 0.2
        if 'EB' in string:
            time += 0.3

        RC[i-1] = time

def SME_RC(i):  # SME的RC計算
    exstring = str(material[i])
    if (process_code[i] == 'SM' and spec[i] != 'nan') or (process_code[i] == 'SM' and 'CRES' in exstring):
        time = 0
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length) - 2 #外放2

        if a < 11:
            time += 0.25
        elif 11 <= a < 21:
            time += 0.27
        elif 21 <= a < 31:
            time += 0.3
        elif 31 <= a < 41:
            time += 0.35
        elif 41 <= a < 51:
            time += 0.4
        elif 51 <= a < 61:
            time += 0.45
        elif 61 <= a < 71:
            time += 0.5
        elif 71 <= a < 81:
            time += 0.55
        elif 81 <= a < 91:
            time += 0.6
        else :
            time += 0.7

        string = str(outsourcing[i])
        if '五面' in string:
            time += 0.3
        if '手臂' in string:
            time += 0.5
        if '倒角' in string:
            time += 0.2
        if '劃切' in string:
            time += 0.2
        if 'EB' in string:
            time += 0.3

        RC[i-1] = time
              
def SC_RC(i):   # SC的RC計算
    if process_code[i] == 'SC':
        time = 0
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length)
        if a < 41:
            time += 1
        elif 41 <= a < 61:
            time += 1.5
        elif 61 <= a < 81:
            time += 2
        elif 81 <= a < 101:
            time += 3
        elif 101 <= a < 121:
            time += 4
        else:
            time += 5

        string = str(outsourcing[i])
        if '五面' in string:
            time += 0.3
        if '手臂' in string:
            time += 0.5

        RC[i-1] = time
            
def SSC_RC(i):  # SSC的RC計算
    if process_code[i] == 'SSC':
        time = 0
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length)
        string = str(outsourcing[i])
        if a >= 100:
            a -= 33
        elif 50 < a < 100:
            a -= 30
            
        if a < 31:
            time += 2
        elif 31 <= a < 51:
            time += 2.5
        elif 51 <= a < 71:
            time += 3
        elif 71 <= a < 91:
            time += 4
        else:
            time += 5
            
        n = time
        n1 = 0
        n2 = 0
        if '拋光' in string: #確認規則後更改
            n1 = float(n) / 2
            time += n1
        if '蝕' in string:
            matches = re.findall(r'\d+', string)
            # 將找到的數字列表轉換為整數列表
            n2 = int(matches[0]) if matches else 0
            n2 *= 0.5
            time += n2
        if '五面' in string:
            time += 0.3
        if '手臂' in string:
            time += 0.2
            
        time += n
        RC[i-1] = str(time)
            
def SSCM_RC(i): # SSCM的RC計算
    if process_code[i] == 'SSCM':
        time = 0
        length = 0
        width = 0
        n1 = 0
        n2 = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length)#外放
        string = str(outsourcing[i])
        m = material[i]

        if 50 <= a < 95:
            a -= 20
        elif a >= 95:
            a -= 30

        if a < 31:
            time += 2
        elif 31 <= a < 51:
            time += 2.5
        elif 51 <= a < 71:
            time += 3
        elif 71 <= a < 91:
            time += 4
        else :
            time += 5
        
        n = time
        if 'POLISHED' in m:
            n1 = float(n) / 2
            time += n1
        if '蝕' in string:
            matches = re.findall(r'\d+', string)
            # 將找到的數字列表轉換為整數列表
            n2 = int(matches[0]) if matches else 0
            n2 *= 0.5
            time += (n2 + n)
        if '五面' in string:
            time += 0.3
        if '手臂' in string:
            time += 0.5

        RC[i-1] = str(time) + " 不含機工"
            
def SSM_RC(i):  # SSM的RC計算
    if process_code[i] == 'SSM':
        time = 0
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length)
        string = str(outsourcing[i])
        
        if a > 50:
            a -= 25
            
        if a < 31:
            time += 2
        elif 31 <= a < 51:
            time += 2.5
        elif 51 <= a < 71:
            time += 3
        elif 71 <= a < 91:
            time += 4
        else:
            time += 5
            
        RC[i-1] = str(time) + '('
        if '機工' in string:
            RC[i-1] += '不含機工'
        if '珠擊' in string:
            RC[i-1] += '/珠擊'
        RC[i-1] += ')'
                
def SSEM_RC(i): # SSEM的RC計算
    if process_code[i] == 'SSEM':
        time = 0
        length = 0
        width = 0
        if not np.isnan(lengthin[i]):
            length = lengthin[i]
        if not np.isnan(widthin[i]):
            width = widthin[i]
        a = max(width, length)#外放
        string = str(outsourcing[i])

        if a > 100:
            a -= 90

        if a < 31:
            time += 2.5
        elif 31 <= a < 51:
            time += 3
        elif 51 <= a < 71:
            time += 4
        elif 71 <= a < 91:
            time += 5
        else :
            time += 6
        

        RC[i-1] = str(time) + "("  + "不含機工" + ")"

read_file()
function()
save_file()